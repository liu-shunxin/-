from openpyxl import Workbook, load_workbook
import pandas as pd
import numpy as np
import math



"""-1"""
# 读取Excel文件，“sheet_name = ”用来指定表名，默认为第一个表
df1 = pd.read_excel("导线网平差作业.xlsx", sheet_name = "观测值和平差值")
df2 = pd.read_excel("导线网平差作业.xlsx", sheet_name = "坐标平差值及误差")

# 将所有非数值项替换为NaN
df1 = df1.apply(pd.to_numeric, errors = "coerce")
df2 = df2.apply(pd.to_numeric, errors = "coerce")

# 提取第一个表第2、3、4、7、8、9列和第二个表第2、3、4、5、6并去除空值
# b_guan为边观测值，b_p为边平差值，b_gai为边改正数，单位：m
# j_guan为角观测值（单位：rad），j_p为角平差值（单位：rad），j_gai为角改正值（单位："）
b_guan = df1.iloc[:, 1].dropna().astype(float).tolist()
b_p = df1.iloc[:, 2].dropna().astype(float).tolist()
b_gai = df1.iloc[:, 3].dropna().astype(float).tolist()
j_guan = df1.iloc[:, 6].dropna().astype(float).tolist()
j_p = df1.iloc[:, 7].dropna().astype(float).tolist()
j_gai = df1.iloc[:, 8].dropna().astype(float).tolist()
# x_p为x坐标平差值，y_p为y坐标平差值，x_g为x坐标改正数，y_g为y坐标改正数，单位：m
# m为点位中误差，单位：m
x_p = df2.iloc[:, 1].dropna().astype(float).tolist()
y_p = df2.iloc[:, 2].dropna().astype(float).tolist()
x_g = df2.iloc[:, 3].dropna().astype(float).tolist()
y_g = df2.iloc[:, 4].dropna().astype(float).tolist()
m = df2.iloc[:, 5].dropna().astype(float).tolist()

# 已知点
x_p[0], y_p[0] = 288.402, 190.56
x_p[1], y_p[1] = 268.08, 289.986



"""0"""
# t为必要观测数
n = len(j_guan) + len(b_guan)

X_p = np.array([x_p[2], y_p[2], x_p[3], y_p[3], x_p[4], y_p[4], x_p[5], y_p[5], x_p[6], y_p[6], x_p[7], y_p[7], x_p[8], y_p[8]])
X_p = X_p.T

t = len(X_p)



"""1、计算待定点近似坐标"""
# α为坐标方位角，a_j为α近似值，单位：rad
a_j = np.zeros((len(x_p), len(x_p)))
a_j[0][1] = math.acos((x_p[1] - x_p[0]) / math.sqrt((x_p[1] - x_p[0]) * (x_p[1] - x_p[0]) + (y_p[1] - y_p[0]) * (y_p[1] - y_p[0])))
a_j[1][0] = a_j[0][1] + math.pi
# 0->6->5->4
a_j[0][6] = a_j[0][1] + j_guan[1]
a_j[6][0] = a_j[0][6] - math.pi
a_j[6][5] = a_j[6][0] - j_guan[9] + 2 * math.pi
a_j[5][6] = a_j[6][5] - math.pi
a_j[5][4] = a_j[5][6] - j_guan[7] + 2 * math.pi
a_j[4][5] = a_j[5][4] + math.pi
# 1->2->3->4
a_j[1][2] = a_j[1][0] + j_guan[0] - 2 * math.pi
a_j[2][1] = a_j[1][2] + math.pi
a_j[2][3] = a_j[2][1] + j_guan[3] - 2 * math.pi
a_j[3][2] = a_j[2][3] - math.pi
a_j[3][4] = a_j[3][2] + j_guan[5]
a_j[4][3] = a_j[3][4] - math.pi
# 1->8->7->5
a_j[1][8] = a_j[1][0] + j_guan[2] - 2 * math.pi
a_j[8][1] = a_j[1][8] - math.pi
a_j[8][7] = a_j[8][1] - j_guan[12] + 2 * math.pi
a_j[7][8] = a_j[8][7] - math.pi
a_j[7][5] = a_j[7][8] - j_guan[11] + 2 * math.pi
a_j[5][7] = a_j[7][5] - math.pi

# x_j为x坐标近似值，y_j为y坐标近似值，单位：m
x_j, y_j = np.zeros(len(x_p)), np.zeros(len(y_p))
x_j[0], y_j[0] = x_p[0], y_p[0]
x_j[1], y_j[1] = x_p[1], y_p[1]
# 0->6->5
x_j[6] = x_j[0] + b_guan[5] * math.cos(a_j[0][6])
y_j[6] = y_j[0] + b_guan[5] * math.sin(a_j[0][6])
x_j[5] = x_j[6] + b_guan[4] * math.cos(a_j[6][5])
y_j[5] = y_j[6] + b_guan[4] * math.sin(a_j[6][5])
# 1->2->3->4
x_j[2] = x_j[1] + b_guan[0] * math.cos(a_j[1][2])
y_j[2] = y_j[1] + b_guan[0] * math.sin(a_j[1][2])
x_j[3] = x_j[2] + b_guan[1] * math.cos(a_j[2][3])
y_j[3] = y_j[2] + b_guan[1] * math.sin(a_j[2][3])
x_j[4] = x_j[3] + b_guan[2] * math.cos(a_j[3][4])
y_j[4] = y_j[3] + b_guan[2] * math.sin(a_j[3][4])
# 1->8->7
x_j[8] = x_j[1] + b_guan[8] * math.cos(a_j[1][8])
y_j[8] = y_j[1] + b_guan[8] * math.sin(a_j[1][8])
x_j[7] = x_j[8] + b_guan[7] * math.cos(a_j[8][7])
y_j[7] = y_j[8] + b_guan[7] * math.sin(a_j[8][7])

# b_j为边近似值（单位：m），l_b = b_guan - b_j（单位：mm）
b_j, l_b = np.zeros(len(b_guan)), np.zeros(len(b_guan))
# b_j
b_j[0] = math.sqrt((x_j[2] - x_j[1]) * (x_j[2] - x_j[1]) + (y_j[2] - y_j[1]) * (y_j[2] - y_j[1]))
b_j[1] = math.sqrt((x_j[3] - x_j[2]) * (x_j[3] - x_j[2]) + (y_j[3] - y_j[2]) * (y_j[3] - y_j[2]))
b_j[2] = math.sqrt((x_j[4] - x_j[3]) * (x_j[4] - x_j[3]) + (y_j[4] - y_j[3]) * (y_j[4] - y_j[3]))
b_j[3] = math.sqrt((x_j[5] - x_j[4]) * (x_j[5] - x_j[4]) + (y_j[5] - y_j[4]) * (y_j[5] - y_j[4]))
b_j[4] = math.sqrt((x_j[6] - x_j[5]) * (x_j[6] - x_j[5]) + (y_j[6] - y_j[5]) * (y_j[6] - y_j[5]))
b_j[5] = math.sqrt((x_j[0] - x_j[6]) * (x_j[0] - x_j[6]) + (y_j[0] - y_j[6]) * (y_j[0] - y_j[6]))
b_j[6] = math.sqrt((x_j[7] - x_j[5]) * (x_j[7] - x_j[5]) + (y_j[7] - y_j[5]) * (y_j[7] - y_j[5]))
b_j[7] = math.sqrt((x_j[8] - x_j[7]) * (x_j[8] - x_j[7]) + (y_j[8] - y_j[7]) * (y_j[8] - y_j[7]))
b_j[8] = math.sqrt((x_j[1] - x_j[8]) * (x_j[1] - x_j[8]) + (y_j[1] - y_j[8]) * (y_j[1] - y_j[8]))
# l_b
for i in range(len(b_guan)):
    l_b[i] = 1000 * (b_guan[i] - b_j[i])



"""2、计算各边坐标方位角改正数方程的系数"""
# a_jiao、b_jiao为坐标方位角改正数方程系数（单位："/mm），p为ρ"，a_为α（单位：rad）
a_jiao, b_jiao, a_ = np.zeros(len(b_guan)), np.zeros(len(b_guan)), np.zeros(len(b_guan))
p = 206797.1649484536
# a_
a_[0] = a_j[1][2]
a_[1] = a_j[2][3]
a_[2] = a_j[3][4]
a_[3] = a_j[5][4]
a_[4] = a_j[6][5]
a_[5] = a_j[0][6]
a_[6] = a_j[7][5]
a_[7] = a_j[8][7]
a_[8] = a_j[1][8]
# a_jiao、b_jiao
for i in range(len(b_guan)):
    a_jiao[i] = p * math.sin(a_[i]) / b_j[i] / 1000
    b_jiao[i] = - p * math.cos(a_[i]) / b_j[i] / 1000



"""3、确定角和边的权"""
# m0为单位权中误差（测角（中）误差）（单位："），m_b为测边误差（单位：1）
m0, m_b = 12, 1 / 2000
# P_jiao为角度观测值的权（单位：1），P_bian为各导线边的权（单位：秒（"）^ 2 / mm ^ 2）
P_jiao = np.ones(len(j_guan))
P_bian = np.zeros(len(b_guan))
for i in range(len(b_guan)):
    P_bian[i] = m0 * m0 / ((1000 * m_b) * (1000 * m_b) * b_guan[i])
# P
P = np.zeros((n, n))
for i in range(len(j_guan)):
    P[i][i] = P_jiao[i]
for i in range(len(b_guan)):
    P[i + len(j_guan)][i + len(j_guan)] = P_bian[i]



"""4、计算角度和边长误差方程系数和常数项"""
# B为系数，l为常数项
B = np.zeros((n, t))
l = np.zeros(n)

# a_bian、b_bain为边长误差方程系数，单位：1
a_bian, b_bian = np.zeros(len(b_guan)), np.zeros(len(b_guan))
# a_bian
a_bian[0] = x_j[2] - x_j[1]
a_bian[1] = x_j[3] - x_j[2]
a_bian[2] = x_j[4] - x_j[3]
a_bian[3] = x_j[5] - x_j[4]
a_bian[4] = x_j[6] - x_j[5]
a_bian[5] = x_j[0] - x_j[6]
a_bian[6] = x_j[7] - x_j[5]
a_bian[7] = x_j[8] - x_j[7]
a_bian[8] = x_j[1] - x_j[8]
for i in range(len(b_guan)):
    a_bian[i] /= b_j[i]
# b_bian
b_bian[0] = y_j[2] - y_j[1]
b_bian[1] = y_j[3] - y_j[2]
b_bian[2] = y_j[4] - y_j[3]
b_bian[3] = y_j[5] - y_j[4]
b_bian[4] = y_j[6] - y_j[5]
b_bian[5] = y_j[0] - y_j[6]
b_bian[6] = y_j[7] - y_j[5]
b_bian[7] = y_j[8] - y_j[7]
b_bian[8] = y_j[1] - y_j[8]
for i in range(len(b_guan)):
    b_bian[i] /= b_j[i]

# B
# j_guan[0]
B[0][0], B[0][1] = - a_jiao[0], - b_jiao[0]
# j_guan[1]
B[1][8], B[1][9] = - a_jiao[5], - b_jiao[5]
# j_guan[2]
B[2][12], B[2][13] = - a_jiao[8], - b_jiao[8]
# j_guan[3]
B[3][0], B[3][1] = a_jiao[1] - a_jiao[0], b_jiao[1] - b_jiao[0]
B[3][2], B[3][3] = - a_jiao[1], - b_jiao[1]
# j_guan[4]
B[4][0], B[4][1] = a_jiao[0], b_jiao[0]
B[4][12], B[4][13] = - a_jiao[8], - b_jiao[8]
# j_guan[5]
B[5][0], B[5][1] = a_jiao[1], b_jiao[1]
B[5][2], B[5][3] = a_jiao[2] - a_jiao[1], b_jiao[2] - b_jiao[1]
B[5][4], B[5][5] = - a_jiao[2], - b_jiao[2]
# j_guan[6]
B[6][2], B[6][3] = a_jiao[2], b_jiao[2]
B[6][4], B[6][5] = a_jiao[3] - a_jiao[2], b_jiao[3] - b_jiao[2]
B[6][6], B[6][7] = - a_jiao[3], - b_jiao[3]
# j_guan[7]
B[7][4], B[7][5] = a_jiao[3], b_jiao[3]
B[7][6], B[7][7] = a_jiao[4] - a_jiao[3], b_jiao[4] - b_jiao[3]
B[7][8], B[7][9] = - a_jiao[4], - b_jiao[4]
# j_guan[8]
B[8][4], B[8][5] = a_jiao[3], b_jiao[3]
B[8][6], B[8][7] = a_jiao[6] - a_jiao[3], b_jiao[6] - b_jiao[3]
B[8][10], B[8][11] = - a_jiao[6], - b_jiao[6]
# j_guan[9]
B[9][6], B[9][7] = a_jiao[4], b_jiao[4]
B[9][8], B[9][9] = a_jiao[5] - a_jiao[4], b_jiao[5] - b_jiao[4]
# j_guan[10]
B[10][6], B[10][7] = a_jiao[6] - a_jiao[4], b_jiao[6] - a_jiao[4]
B[10][8], B[10][9] = a_jiao[4], b_jiao[4]
B[10][10], B[10][11] = - a_jiao[6], - b_jiao[6]
# j_guan[11]
B[11][6], B[11][7] = a_jiao[6], b_jiao[6]
B[11][10], B[11][11] = a_jiao[7] - a_jiao[6], b_jiao[7] - b_jiao[6]
B[11][12], B[11][13] = - a_jiao[8], - b_jiao[8]
# j_guan[12]
B[12][10], B[12][11] = a_jiao[7], b_jiao[8]
B[12][12], B[12][13] = a_jiao[8] - a_jiao[7], b_jiao[8] - b_jiao[7]
# b_guan[0]
B[13][0], B[13][1] = a_bian[0], b_bian[0]
# b_guan[1]
B[14][0], B[14][1] = - a_bian[1], - b_bian[1]
B[14][2], B[14][3] = a_bian[1], b_bian[1]
# b_guan[2]
B[15][2], B[15][3] = - a_bian[2], - b_bian[2]
B[15][4], B[15][5] = a_bian[2], b_bian[2]
# b_guan[3]
B[16][4], B[16][5] = - a_bian[3], - b_bian[3]
B[16][6], B[16][7] = a_bian[3], b_bian[3]
# b_guan[4]
B[17][6], B[17][7] = - a_bian[4], - b_bian[4]
B[17][8], B[17][9] = a_bian[4], b_bian[4]
# b_guan[5]
B[18][8], B[18][9] = - a_bian[5], - b_bian[5]
# b_guan[6]
B[19][6], B[19][7] = - a_bian[6], - b_bian[6]
B[19][10], B[19][11] = a_bian[6], b_bian[6]
# b_guan[7]
B[20][10], B[20][11] = - a_bian[7], - b_bian[7]
B[20][12], B[20][13] = a_bian[7], b_bian[7]
# b_guan[8]
B[21][12], B[21][13] = - a_bian[8], - b_bian[8]

# l
l = np.zeros(n)
# j_j为角近似值，单位："
j_j = np.zeros(len(j_guan))
j_j[0] = j_guan[0]
j_j[1] = j_guan[1]
j_j[2] = j_guan[2]
j_j[3] = j_guan[3]
j_j[4] = a_j[1][8] - a_j[1][2]
j_j[5] = j_guan[5]
j_j[6] = a_j[4][5] - a_j[4][3]
j_j[7] = j_guan[7]
j_j[8] = a_j[5][7] - a_j[5][4] + 2 * math.pi
j_j[9] = j_guan[9]
j_j[10] = a_j[5][7] - a_j[5][6]
j_j[11] = j_guan[11]
j_j[12] = j_guan[12]
# 角
for i in range(len(j_guan)):
    l[i] = (j_guan[i] - j_j[i]) / math.pi * 180 * 3600
# 边
for i in range(len(b_guan)):
    l[len(j_guan) + i] = b_guan[i] - b_j[i]



"""5、法方程的组成和解算"""
# N为法方程的系数项，B.T @ P @ l为常数项
N = B.T @ P @ B
# x_gai为坐标（参数）改正数
x_gai = np.linalg.inv(N) @ B.T @ P @ l
# v为角度和边长改正数
v = B @ x_gai - l



"""6、平差值计算"""
# （1）坐标平差值
for i in range(len(x_p) - 2):
    # x_p
    x_p[i + 2] = x_j[i + 2] + x_gai[2 * i]
    # y_p
    y_p[i + 2] = y_j[i + 2] + x_gai[2 * i + 1]
# （2）观测值平差值
# j_p
for i in range(len(j_guan)):
    j_p[i] = j_guan[i] + v[i] / 3600 / 180 * math.pi
# b_p
b_p[0] = math.sqrt((x_p[1] - x_p[0]) * (x_p[1] - x_p[0]) + (y_p[1] - y_p[0]) * (y_p[1] - y_p[0]))
for i in range(len(b_guan)):
    b_p[i + 1] = b_guan[i] + v[i + len(j_guan)]



"""7、精度计算"""
# （1）单位权中误差（测角中误差）
m0_p = math.sqrt(v.T @ P @ v / (n - t))
# （2）待定点点位中误差
# Q为未知数的权倒数，单位：mm ^ 2 / "（秒） ^ 2
Q = np.linalg.inv(N)
m = np.zeros(len(x_j))
for i in range(len(x_p) - 2):
    m[i + 2] = m0_p * math.sqrt(Q[2 * i][2 * i] + Q[2 * i + 1][2 * i + 1])
    


"""写入Excel文件"""
# b_gai
for i in range(len(b_guan)):
    b_gai[i + 1] = v[len(j_guan) + i]
# j_gai
for i in range(len(j_guan)):
    j_gai[i] = v[i]
# x_g
for i in range(len(x_p) - 2):
    x_g[i + 2] = x_gai[2 * i]
# y_g
for i in range(len(y_p) - 2):
    y_g[i + 2] = x_gai[2 * i + 1]

# 定义写入数据
data = {
    "观测值和平差值": {
        3: b_p, 4: b_gai, 8: j_p, 9: j_gai,
    },
    "坐标平差值及误差": {
        2: x_p, 3: y_p, 4: x_g, 5: y_g, 6: m,
    },
}

# 尝试加载Excel文件，不存在是创建一个新文件
try:
    workbook = load_workbook("导线网平差作业.xlsx")
except FileNotFoundError:
    workbook = Workbook()

# 遍历数据并写入到不同的工作表和列
for sheet_name, columns in data.items():
    # 检查工作表是否存在，如果不存在则创建
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]

    # 遍历指定的列和数据
    for column, values in columns.items():
    # 写入每列的数据到指定列（覆盖原有数据）
        for row, value in enumerate(values, start = 2):
            sheet.cell(row = row, column = column, value = value)

workbook.save("导线网平差作业.xlsx")



""""""
# 平差后最弱边相对中误差
m_b_max = 0
for i in range(len(b_guan)):
    if m_b_max < abs(b_gai[i + 1] / b_guan[i]):
        m_b_max = abs(b_gai[i + 1] / b_guan[i])
print(m_b_max)
# 平差后最弱点点位中误差（mm）
m_max = 0
for i in range(len(m)):
    if m_max < m[i]:
        m_max = m[i]
print(m_max)
# 平差后单位权中误差（"）
print(m0_p)